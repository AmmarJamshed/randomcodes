#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pakistan Adverse Media Scraper - portable single file
====================================================
Copy this file AND keywords.xlsx to any PC. PDFs saved as:
  downloads/<Keyword>/<Source>/<YYYY-MM-DD>/<title>.pdf

Keyword lists live in keywords.xlsx (edit in Excel; restart after changes).

SETUP (once per machine):
  pip install requests beautifulsoup4 reportlab deep-translator lxml openpyxl

RUN (one edition at a time):
  python scraper.py
      Prompt: 23-07-2026 14:00
      Scrapes ALL sources for that date, then exits.
      Re-run and enter the next date/time (e.g. Sat, then Sun, then Mon).

  python scraper.py --date "23-07-2026 14:00"
  python scraper.py --sources Dawn "The News"
  python scraper.py --fresh
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import shutil
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import urljoin, urlparse, unquote

import requests
from bs4 import BeautifulSoup
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None  # type: ignore

try:
    from deep_translator import GoogleTranslator
except ImportError:
    GoogleTranslator = None  # type: ignore

BASE = Path(__file__).resolve().parent
OUT = BASE / "downloads"
MANIFEST = BASE / "manifest.csv"
TARGET = date.today().isoformat()
KEYWORDS_XLSX = BASE / "keywords.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}
TIMEOUT = 25
DELAY = 0.8


def _load_keywords() -> Tuple[Dict[str, List[str]], List[str]]:
    if load_workbook is None:
        raise SystemExit("Install openpyxl to read keywords.xlsx:  pip install openpyxl")
    path = KEYWORDS_XLSX
    if not path.exists():
        raise SystemExit(
            f"Missing {path.name} next to scraper.py — copy keywords.xlsx with this script."
        )
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise SystemExit(f"Could not open keywords.xlsx: {e}") from e

    if "keywords" not in wb.sheetnames:
        raise SystemExit("keywords.xlsx must have a sheet named 'keywords'.")

    kws: Dict[str, List[str]] = {}
    seen: Dict[str, Set[str]] = {}
    ws = wb["keywords"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    for row in rows:
        if not row or len(row) < 2:
            continue
        cat, term = row[0], row[1]
        if cat is None or term is None:
            continue
        cat_s, term_s = str(cat).strip(), str(term).strip()
        if not cat_s or not term_s or cat_s.lower() == "category":
            continue
        bucket = seen.setdefault(cat_s, set())
        if term_s.lower() in {t.lower() for t in bucket}:
            continue
        bucket.add(term_s)
        kws.setdefault(cat_s, []).append(term_s)

    pri: List[str] = []
    if "priority" in wb.sheetnames:
        pr = wb["priority"]
        prows = pr.iter_rows(values_only=True)
        next(prows, None)
        for row in prows:
            if not row:
                continue
            # accept (order, category) or (category,)
            cat = row[1] if len(row) > 1 and row[1] is not None else row[0]
            if cat is None:
                continue
            cat_s = str(cat).strip()
            if not cat_s or cat_s.lower() in {"category", "order"}:
                continue
            if cat_s not in pri:
                pri.append(cat_s)
    wb.close()

    if not kws:
        raise SystemExit("keywords.xlsx sheet 'keywords' has no category/term rows.")
    if not pri:
        pri = list(kws.keys())
    return kws, pri


KEYWORDS, PRIORITY = _load_keywords()

# Places named in the manifest `city` column (not used to skip articles).
# Longer names first so "South Waziristan" wins over "Waziristan".
PLACE_NAMES = [
    "South Waziristan", "North Waziristan", "Dera Ismail Khan", "Dera Ghazi Khan",
    "Dera Bugti", "Rahim Yar Khan", "Sheikhupura", "Mandi Bahauddin",
    "Tando Muhammad Khan", "Tando Allahyar", "Shaheed Benazirabad",
    "Naushahro Feroze", "Qilla Saifullah", "Qilla Abdullah", "Qila Saifullah",
    "Qila Abdullah", "Khyber Pakhtunkhwa", "Gilgit-Baltistan", "Azad Kashmir",
    "Azad Jammu", "Islamabad", "Rawalpindi", "Faisalabad", "Gujranwala",
    "Bahawalpur", "Sargodha", "Abbottabad", "Muzaffarabad", "Hyderabad",
    "Nawabshah", "Jacobabad", "Shikarpur", "Khairpur", "Jamshoro",
    "Tharparkar", "Umerkot", "Sanghar", "Kashmore", "Ghotki", "Larkana",
    "Karachi", "Lahore", "Peshawar", "Quetta", "Multan", "Sukkur", "Gwadar",
    "Sialkot", "Gujrat", "Sahiwal", "Okara", "Jhang", "Kasur", "Attock",
    "Chakwal", "Jhelum", "Hafizabad", "Narowal", "Pakpattan", "Vehari",
    "Khanewal", "Lodhran", "Layyah", "Bhakkar", "Mianwali", "Khushab",
    "Chiniot", "Mardan", "Mingora", "Charsadda", "Nowshera", "Mansehra",
    "Haripur", "Malakand", "Shangla", "Battagram", "Kohistan", "Torghar",
    "Lakki Marwat", "Hangu", "Kohat", "Bannu", "Bajaur", "Mohmand",
    "Orakzai", "Kurram", "Waziristan", "Hassankhel", "Hassan Khel",
    "Upper Dir", "Lower Dir", "Chitral", "Swabi", "Buner", "Tank",
    "Kachhi", "Kacchi", "Kachi", "Bolan", "Nasirabad", "Jaffarabad",
    "Jhal Magsi", "Kalat", "Mastung", "Khuzdar", "Lasbela", "Awaran",
    "Panjgur", "Washuk", "Kharan", "Nushki", "Chagai", "Pishin", "Zhob",
    "Sherani", "Musakhel", "Loralai", "Harnai", "Ziarat", "Kohlu",
    "Barkhan", "Dalbandin", "Turbat", "Pasni", "Ormara", "Wadh", "Sibi",
    "Thatta", "Badin", "Dadu", "Matiari", "Qambar", "Shahdadkot",
    "North Nazimabad", "Manghopir", "Sohrab Goth", "Ferozwala", "Dir",
    "Swat", "Khyber", "Balochistan", "Sindh", "Punjab", "Gilgit",
    "Mirpur", "AJK", "ICT", "Pindi", "DI Khan", "DG Khan",
    "Pakistan",
]
_PLACE_RE = re.compile(
    r"\b("
    + "|".join(re.escape(p) for p in sorted(set(PLACE_NAMES), key=len, reverse=True))
    + r")\b",
    re.I,
)
PLACE_CANONICAL = {
    "kacchi": "Kachhi",
    "kachi": "Kachhi",
    "kachhi": "Kachhi",
    "pindi": "Rawalpindi",
    "hassankhel": "Hassan Khel",
    "hassan khel": "Hassan Khel",
    "di khan": "Dera Ismail Khan",
    "dg khan": "Dera Ghazi Khan",
    "ict": "Islamabad",
    "ajk": "Azad Kashmir",
}

EXCLUDE = re.compile(
    r"(cricket|football|bollywood|hollywood|movie|celebrity|fashion|editorial|"
    r"stock market|KSE-100|monsoon spell|climate crisis|world cup)",
    re.I,
)

MANIFEST_FIELDS = [
    "record_id", "title", "url", "keyword", "all_keywords",
    "source", "city", "date", "pdf_path", "pdf_link", "person_names", "scraped_at",
]

session = requests.Session()
session.headers.update(HEADERS)


def pdf_hyperlink(pdf: Path) -> str:
    """Excel-clickable HYPERLINK formula pointing at the saved PDF."""
    uri = pdf.resolve().as_uri()  # file:///D:/...
    uri_esc = uri.replace('"', '""')
    return f'=HYPERLINK("{uri_esc}","Open PDF")'


def resolve_pdf_from_manifest(pdf_path_val: str) -> Optional[Path]:
    """Resolve a relative path, absolute path, or file URI from a manifest row."""
    if not pdf_path_val:
        return None
    raw = pdf_path_val.strip().strip('"')
    m = re.search(r'HYPERLINK\("([^"]+)"', raw, re.I)
    if m:
        raw = m.group(1)
    if raw.lower().startswith("file:"):
        parsed = urlparse(raw)
        path = unquote(parsed.path)
        if os.name == "nt" and path.startswith("/") and len(path) >= 3 and path[2] == ":":
            path = path[1:]
        return Path(path)
    p = Path(raw)
    if p.is_absolute():
        return p
    cand = OUT / p
    if cand.exists():
        return cand
    cand2 = BASE / p
    return cand2 if cand2.exists() else cand


def sync_manifest_xlsx(csv_path: Path) -> None:
    """Write/refresh a .xlsx next to the CSV with real clickable PDF hyperlinks."""
    if load_workbook is None or not csv_path.exists():
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return
    try:
        with csv_path.open(encoding="utf-8", newline="") as f:
            rows = list(csv.DictReader(f))
            fields = list(rows[0].keys()) if rows else list(MANIFEST_FIELDS)
    except OSError:
        return
    if "pdf_link" not in fields:
        # insert after pdf_path if present
        if "pdf_path" in fields:
            i = fields.index("pdf_path") + 1
            fields = fields[:i] + ["pdf_link"] + [c for c in fields[i:] if c != "pdf_link"]
        else:
            fields = fields + ["pdf_link"]

    wb = Workbook()
    ws = wb.active
    ws.title = "manifest"
    ws.append(fields)
    link_font = Font(color="0563C1", underline="single")
    for row in rows:
        values = []
        for col in fields:
            if col == "pdf_link":
                values.append("")  # filled as hyperlink below
            else:
                values.append(row.get(col, ""))
        ws.append(values)
        r_idx = ws.max_row
        pdf = resolve_pdf_from_manifest(row.get("pdf_path", "") or "")
        # also parse formula in pdf_link column if present
        if pdf is None and row.get("pdf_link"):
            pdf = resolve_pdf_from_manifest(row.get("pdf_link", ""))
        cell = ws.cell(row=r_idx, column=fields.index("pdf_link") + 1)
        if pdf and pdf.exists():
            cell.hyperlink = pdf.resolve().as_uri()
            cell.value = "Open PDF"
            cell.font = link_font
        elif row.get("pdf_link", "").startswith("="):
            cell.value = row["pdf_link"]
        else:
            cell.value = row.get("pdf_link", "") or ""
    xlsx_path = csv_path.with_suffix(".xlsx")
    try:
        wb.save(xlsx_path)
    except OSError:
        pass


def get(url: str, referer: str = "") -> Optional[BeautifulSoup]:
    try:
        h = {"Referer": referer} if referer else {}
        r = session.get(url, headers=h, timeout=TIMEOUT)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        time.sleep(DELAY)
        return BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"  [WARN] {url}: {e}")
        return None


def translate(text: str) -> str:
    if not text or not text.strip() or GoogleTranslator is None:
        return (text or "").strip()
    letters = [c for c in text if c.isalpha()]
    if letters and sum(1 for c in letters if ord(c) < 128) / len(letters) > 0.7:
        return text.strip()
    try:
        out = []
        chunk = text.strip()
        for i in range(0, len(chunk), 4500):
            part = chunk[i : i + 4500]
            out.append(GoogleTranslator(source="ur", target="en").translate(part) or part)
            time.sleep(0.25)
        return " ".join(out).strip()
    except Exception:
        return text.strip()


def find_keywords(title: str) -> List[str]:
    hits = []
    for folder, terms in KEYWORDS.items():
        pat = re.compile(
            r"\b(?:" + "|".join(re.escape(t) for t in terms) + r")\b", re.I
        )
        if pat.search(title or ""):
            hits.append(folder)
    return hits


def extract_city(title: str, body: str = "", url: str = "") -> str:
    """Name places found in the story for the manifest; never used to skip."""
    found: List[str] = []
    seen: Set[str] = set()
    blob = f"{title or ''} {body or ''} {url or ''}"
    for m in _PLACE_RE.finditer(blob):
        raw = m.group(1)
        name = PLACE_CANONICAL.get(raw.lower(), raw)
        key = name.lower()
        if key in seen or key == "pakistan":
            continue
        seen.add(key)
        found.append(name)
    return "; ".join(found)


def is_relevant(title: str, body: str) -> Tuple[bool, List[str], str]:
    """Save every keyword hit. Location is recorded in `city`, not used to filter."""
    kws = find_keywords(title)
    if not kws:
        return False, [], "no_keyword"
    if EXCLUDE.search(title or ""):
        return False, [], "excluded"
    return True, kws, "ok"


def primary_kw(kws: List[str]) -> str:
    for p in PRIORITY:
        if p in kws:
            return p
    return kws[0] if kws else "General"


def story_id(title: str) -> str:
    t = re.sub(r"[^\w\s]", " ", (title or "").lower())
    t = re.sub(r"\s+", " ", t).strip()
    return hashlib.md5(t.encode()).hexdigest()[:12]


def safe_name(title: str) -> str:
    t = re.sub(r'[<>:"/\\|?*]', "", title)
    t = re.sub(r"\s+", "_", t.strip())[:80]
    return t or "article"


def save_pdf(
    path: Path, title: str, source: str, url: str, date_str: str, body: str, kws: List[str]
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle(
        "T", parent=styles["Heading1"], fontSize=14, leading=18,
        alignment=TA_LEFT, spaceAfter=10,
    )
    meta_s = ParagraphStyle(
        "M", parent=styles["Normal"], fontSize=9, textColor="#444", spaceAfter=4,
    )
    body_s = ParagraphStyle(
        "B", parent=styles["Normal"], fontSize=11, leading=15,
        alignment=TA_JUSTIFY, spaceAfter=6,
    )

    def esc(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    story = [
        Paragraph(esc(title), title_s),
        Spacer(1, 0.08 * inch),
        Paragraph(f"<b>Source:</b> {esc(source)}", meta_s),
        Paragraph(f"<b>Date:</b> {esc(date_str)}", meta_s),
        Paragraph(f"<b>URL:</b> {esc(url)}", meta_s),
        Paragraph(f"<b>Keywords:</b> {esc(', '.join(kws))}", meta_s),
        Spacer(1, 0.15 * inch),
    ]
    for para in re.split(r"(?<=[.!?])\s+(?=[A-Z])", (body or "").strip()) or [body]:
        if para and para.strip():
            story.append(Paragraph(esc(para.strip()), body_s))
    SimpleDocTemplate(
        str(path), pagesize=LETTER,
        rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50,
    ).build(story)


def load_seen() -> Set[str]:
    """Duplicate check uses title hashes from manifest CSV files — NOT downloads/."""
    seen: Set[str] = set()
    for p in [MANIFEST, *BASE.glob("manifest_*.csv")]:
        if not p.exists():
            continue
        try:
            with p.open(encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("title"):
                        seen.add(story_id(row["title"]))
        except OSError:
            pass
    return seen


def clear_manifests() -> None:
    """Delete all manifest files used for duplicate detection."""
    for p in [MANIFEST, MANIFEST.with_suffix(".xlsx"), *BASE.glob("manifest_*.csv"), *BASE.glob("manifest_*.xlsx")]:
        # keep keywords.xlsx
        if p.name.lower() == "keywords.xlsx":
            continue
        try:
            if p.exists():
                p.unlink()
        except OSError:
            pass


def append_manifest(row: dict) -> None:
    payload = {k: row.get(k, "") for k in MANIFEST_FIELDS}
    if not payload.get("pdf_link") and row.get("pdf_abs"):
        payload["pdf_link"] = pdf_hyperlink(Path(row["pdf_abs"]))
    for path in [MANIFEST, BASE / f"manifest_{row['date']}.csv"]:
        try:
            ensure_manifest_schema(path)
            exists = path.exists()
            with path.open("a", encoding="utf-8", newline="") as f:
                w = csv.DictWriter(f, fieldnames=MANIFEST_FIELDS)
                if not exists:
                    w.writeheader()
                w.writerow(payload)
            sync_manifest_xlsx(path)
        except PermissionError:
            continue


def ensure_manifest_schema(path: Path) -> None:
    """Add city / pdf_link columns to older manifest files without dropping rows."""
    if not path.exists():
        return
    try:
        with path.open(encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            old_fields = list(reader.fieldnames or [])
            rows = list(reader)
    except OSError:
        return
    # Preserve any extra columns from older/mixed writers
    fields = list(MANIFEST_FIELDS)
    for f in old_fields:
        if f and f not in fields:
            fields.append(f)
    needs_rewrite = old_fields != fields
    upgraded = []
    for row in rows:
        city = (row.get("city") or "").strip() or extract_city(
            row.get("title", ""), "", row.get("url", "")
        )
        pdf_path = row.get("pdf_path", "") or ""
        pdf_link = (row.get("pdf_link") or "").strip()
        if not pdf_link and pdf_path:
            pdf = resolve_pdf_from_manifest(pdf_path)
            if pdf:
                pdf_link = pdf_hyperlink(pdf)
                needs_rewrite = True
        if not (row.get("city") or "").strip() and city:
            needs_rewrite = True
        if "pdf_link" not in old_fields:
            needs_rewrite = True
        out = {k: row.get(k, "") for k in fields}
        out["city"] = city
        out["pdf_link"] = pdf_link
        upgraded.append(out)
    if not needs_rewrite:
        sync_manifest_xlsx(path)
        return
    try:
        tmp = path.with_suffix(path.suffix + ".tmp")
        with tmp.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fields)
            w.writeheader()
            w.writerows(upgraded)
        tmp.replace(path)
        sync_manifest_xlsx(path)
    except OSError:
        return


def scrape_dawn(target: str) -> List[dict]:
    print("=== Dawn ===")
    dt = datetime.strptime(target, "%Y-%m-%d")
    date_key = dt.strftime("%d_%m_%Y")
    edition = f"{dt.year}/{dt.month:02d}/{dt.day:02d}"
    listing = get(f"https://epaper.dawn.com/?page={date_key}_001", "https://epaper.dawn.com/")
    html = str(listing) if listing else ""
    page_ids = sorted(set(re.findall(rf"[?&]page=({re.escape(date_key)}_\d{{3}})", html, flags=re.I)))
    if not page_ids:
        page_ids = [f"{date_key}_001"]
    print(f"  Listed pages from HTML: {len(page_ids)}")

    codes: Set[str] = set()
    for page_id in page_ids:
        page_url = f"https://e.dawn.com/{edition}/pages/{page_id}.html"
        soup = get(page_url, "https://epaper.dawn.com/")
        page_html = str(soup) if soup else ""
        if not soup or len(page_html) < 400:
            alt = get(f"https://epaper.dawn.com/?page={page_id}", "https://epaper.dawn.com/")
            page_html = str(alt) if alt else page_html
        found = re.findall(r"StoryText=([^'\"&\s<>)]+)", page_html, flags=re.I)
        found += re.findall(r"StoryImage=([^'\"&\s<>)]+)", page_html, flags=re.I)
        codes.update(c.strip() for c in found if c.strip())

    print(f"  Story codes: {len(codes)}")
    arts = []
    for code in sorted(codes):
        url = f"https://epaper.dawn.com/DetailNews.php?StoryText={code}"
        soup = get(url, "https://epaper.dawn.com/")
        if not soup:
            continue
        title = ""
        if soup.title:
            parts = [p.strip() for p in soup.title.get_text(strip=True).split("|") if p.strip()]
            title = parts[-1] if len(parts) >= 3 else (parts[0] if parts else "")
        if not title or title.lower().startswith("dawn-epaper"):
            for sel in ("h1", "h2"):
                el = soup.select_one(sel)
                if el and el.get_text(strip=True) and "epaper" not in el.get_text(strip=True).lower():
                    title = el.get_text(strip=True)
                    break
            if not title or title.lower().startswith("dawn-epaper"):
                og = soup.find("meta", property="og:title")
                if og and og.get("content"):
                    title = og["content"]
        body = ""
        for sel in ("article", ".story", "#story", ".detail-text"):
            el = soup.select_one(sel)
            if el and len(el.get_text(" ", strip=True)) > 80:
                body = el.get_text(" ", strip=True)
                break
        if not body:
            body = soup.get_text(" ", strip=True)[:3000]
        arts.append({
            "title": title, "url": url, "source": "Dawn", "date": target, "body": body,
        })
    print(f"  Found {len(arts)}")
    return arts


def scrape_thenews(target: str) -> List[dict]:
    print("=== The News ===")
    # Dated city editions: /{city}/{DD-MM-YYYY}/pageN
    # Stories are image-map <area href=".../detail?id=..."> with no visible title.
    dt = datetime.strptime(target, "%Y-%m-%d")
    slug = dt.strftime("%d-%m-%Y")
    cities = ("karachi", "lahore", "pindi")
    ids: Set[str] = set()
    for city in cities:
        first_url = f"https://e.thenews.pk/{city}/{slug}/page1"
        first = get(first_url, "https://e.thenews.pk/")
        if not first or len(str(first)) < 2000:
            print(f"  {city}: no edition HTML")
            continue
        html1 = str(first)
        page_nums = sorted({
            int(n) for n in re.findall(rf"/{re.escape(city)}/{re.escape(slug)}/page(\d+)", html1)
        })
        if not page_nums:
            page_nums = [1]
        print(f"  {city} listed pages: {len(page_nums)} ({min(page_nums)}-{max(page_nums)})")
        for n in page_nums:
            soup = first if n == 1 else get(
                f"https://e.thenews.pk/{city}/{slug}/page{n}", first_url
            )
            if not soup:
                continue
            html = str(soup)
            ids.update(re.findall(r"(?:e\.thenews\.pk)?/detail/?\?id=(\d+)", html, re.I))
            for tag in soup.find_all(["area", "a"], href=True):
                m = re.search(r"[?&]id=(\d+)", tag["href"])
                if m:
                    ids.add(m.group(1))
    print(f"  Detail IDs: {len(ids)}")
    if not ids:
        print("  No archive pages found")
        return []

    out: List[dict] = []
    for i, did in enumerate(sorted(ids, key=int), 1):
        if i % 50 == 0:
            print(f"  Enriched {i}/{len(ids)}...")
        url = f"https://e.thenews.pk/detail/?id={did}"
        s = get(url, "https://e.thenews.pk/")
        if not s:
            continue
        title = ""
        h1 = s.find("h1")
        if h1:
            title = h1.get_text(strip=True)
        if not title:
            og = s.find("meta", property="og:title")
            if og and og.get("content"):
                title = og["content"].strip()
        if not title and s.title:
            title = s.title.get_text(strip=True).split("|")[0].strip()
        if not find_keywords(title):
            continue
        body = ""
        el = s.select_one(".story-detail")
        if el:
            body = el.get_text(" ", strip=True)
        out.append({
            "title": title, "url": url, "source": "The News",
            "date": target, "body": body,
        })
    print(f"  Found {len(out)} keyword hits")
    return out


def scrape_tribune(target: str) -> List[dict]:
    print("=== Express Tribune ===")
    # Dated e-paper + print edition. /today-paper is always the live site.
    epaper_url = f"https://tribune.com.pk/epaper/{target}"
    paper_url = f"https://tribune.com.pk/newspaper/{target}"
    soup = get(epaper_url, "https://tribune.com.pk/epaper")
    clip_paths = []
    if soup:
        clip_paths = sorted(set(re.findall(
            rf"/epaper/news/[A-Za-z]+/{re.escape(target)}/[A-Za-z0-9+/=]+",
            str(soup),
        )))
    print(f"  E-paper clips for {target}: {len(clip_paths)}")

    paper = get(paper_url, "https://tribune.com.pk/epaper")
    if not paper:
        paper = get("https://tribune.com.pk/today-paper", "https://tribune.com.pk/epaper")
        if target != date.today().isoformat():
            print("  [WARN] newspaper archive missing; fell back to today-paper")
    if not paper:
        return []
    skip = (
        "sports", "life-style", "tmagazine", "blogs", "technology",
        "multimedia", "cartoon", "classified",
    )
    seen, arts = set(), []
    for a in paper.find_all("a", href=True):
        href = a["href"]
        if "/story/" not in href or any(x in href.lower() for x in skip):
            continue
        url = href if href.startswith("http") else urljoin("https://tribune.com.pk/", href)
        title = a.get_text(strip=True)
        if url in seen or not title or len(title) < 10:
            continue
        if not find_keywords(title):
            continue
        seen.add(url)
        s = get(url, paper_url)
        body = ""
        if s:
            h1 = s.find("h1")
            if h1:
                title = h1.get_text(strip=True) or title
            for sel in (".story-text", ".story-content", "article"):
                el = s.select_one(sel)
                if el and len(el.get_text(" ", strip=True)) > 80:
                    body = el.get_text(" ", strip=True)
                    break
        arts.append({
            "title": title, "url": url, "source": "Express Tribune",
            "date": target, "body": body,
        })
    print(f"  Found {len(arts)} keyword hits")
    return arts


def scrape_brecorder(target: str) -> List[dict]:
    print("=== Business Recorder ===")
    dt = datetime.strptime(target, "%Y-%m-%d")
    listing = f"https://epaper.brecorder.com/{dt.year}/{dt.month:02d}/{dt.day:02d}/"
    soup = get(listing, "https://epaper.brecorder.com/")
    if not soup:
        return []
    html = str(soup)
    pages = set()
    for a in soup.find_all("a", href=True):
        if re.search(r"/\d{4}/\d{2}/\d{2}/[^\"']+-page\.html", a["href"]):
            pages.add(urljoin(listing, a["href"]))
    pages.update(
        urljoin(listing, p)
        for p in re.findall(r"(/\d{4}/\d{2}/\d{2}/[^\"'\s]+-page\.html)", html)
    )
    if not pages:
        pages.add(listing)
    print(f"  Listed pages from HTML: {len(pages)}")
    urls: List[str] = []
    for page_url in sorted(pages):
        ps = soup if page_url.rstrip("/") == listing.rstrip("/") else get(page_url, listing)
        if not ps:
            continue
        page_html = str(ps)
        for rect in ps.find_all(attrs={"type": "news"}):
            aid, ed, pg = rect.get("id"), rect.get("date"), rect.get("page")
            if aid and ed and pg:
                urls.append(urljoin(page_url, f"/{ed}/{pg}-page/{aid}-news.html"))
        urls.extend(
            urljoin(page_url, m)
            for m in re.findall(
                r"(/\d{4}/\d{2}/\d{2}/[^\"'\s]+-page/\d+-news\.html)", page_html
            )
        )
    print(f"  News clips: {len(set(urls))}")
    arts = []
    for url in sorted(set(urls)):
        s = get(url, listing)
        if not s:
            continue
        title = ""
        h2 = s.find("h2")
        if h2:
            title = h2.get_text(strip=True)
        elif s.title:
            title = s.title.get_text(strip=True).split("|")[0].strip()
        body = ""
        el = s.find("article")
        if el:
            body = el.get_text(" ", strip=True)
        if not find_keywords(title):
            continue
        arts.append({
            "title": title, "url": url, "source": "Business Recorder",
            "date": target, "body": body,
        })
    print(f"  Found {len(arts)} keyword hits")
    return arts


def scrape_jang(target: str) -> List[dict]:
    print("=== Jang (Urdu->English) ===")
    if GoogleTranslator is None:
        print("  [WARN] deep-translator not installed - skipping Jang")
        return []
    get("https://e.jang.com.pk/")
    slug = datetime.strptime(target, "%Y-%m-%d").strftime("%d-%m-%Y")
    ids: Set[str] = set()
    for city in ("karachi", "lahore", "pindi", "multan", "quetta"):
        first = get(f"https://e.jang.com.pk/{city}/{slug}/page1", "https://e.jang.com.pk/")
        if not first or len(str(first)) < 2000:
            print(f"  {city}: no edition")
            continue
        html1 = str(first)
        page_nums = sorted({int(n) for n in re.findall(rf"/{city}/{re.escape(slug)}/page(\d+)", html1)})
        if not page_nums:
            page_nums = list(range(1, 25))
        print(f"  {city} listed pages: {page_nums}")
        for n in page_nums:
            soup = first if n == 1 else get(
                f"https://e.jang.com.pk/{city}/{slug}/page{n}", "https://e.jang.com.pk/"
            )
            if not soup or len(str(soup)) < 500:
                continue
            ids.update(re.findall(r"detail\?id=(\d+)", str(soup)))
            ids.update(re.findall(r"/detail/(\d+)", str(soup)))
    print(f"  Detail IDs: {len(ids)}")
    prefix = re.compile(
        r"^Jang Epaper (?:Karachi|Lahore|Rawalpindi|Multan|Quetta|London)\s*", re.I
    )
    arts = []
    for i, did in enumerate(sorted(ids, key=int), 1):
        if i % 50 == 0:
            print(f"  Enriched {i}/{len(ids)}...")
        url = f"https://e.jang.com.pk/detail/{did}"
        soup = get(url, "https://e.jang.com.pk/")
        if not soup:
            continue
        raw = ""
        if soup.title:
            raw = soup.title.get_text(strip=True)
        og = soup.find("meta", property="og:title")
        if og and og.get("content"):
            raw = og["content"]
        urdu_title = prefix.sub("", raw).strip() or raw
        body = ""
        for sel in (".detail-content", ".detail-view-content", "article"):
            el = soup.select_one(sel)
            if el and len(el.get_text(" ", strip=True)) > 80:
                body = el.get_text(" ", strip=True)
                break
        en_title = translate(urdu_title)
        if not find_keywords(en_title):
            continue
        en_body = translate(body) if body else ""
        arts.append({
            "title": en_title or urdu_title, "url": url,
            "source": "Jang News English", "date": target, "body": en_body,
        })
    print(f"  Found {len(arts)} keyword hits")
    return arts


def scrape_customnews(target: str) -> List[dict]:
    print("=== Custom News ===")
    dt = datetime.strptime(target, "%Y-%m-%d")
    month_url = f"https://customnews.pk/{dt.year}/{dt.month:02d}/"
    day_url = f"https://customnews.pk/{dt.year}/{dt.month:02d}/{dt.day:02d}/"
    urls_try = [
        "https://customnews.pk/",
        month_url,
        day_url,
    ]
    for n in range(2, 12):
        urls_try.append(f"{month_url}page/{n}/")
        urls_try.append(f"https://customnews.pk/page/{n}/")
    seen, arts = set(), []
    for page in urls_try:
        soup = get(page, "https://customnews.pk/")
        if not soup:
            continue
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if not re.search(r"customnews\.pk/\d{4}/\d{2}/\d{2}/", href):
                continue
            url = href if href.startswith("http") else urljoin(page, href)
            m = re.search(r"/(\d{4})/(\d{2})/(\d{2})/", url)
            if not m or f"{m.group(1)}-{m.group(2)}-{m.group(3)}" != target:
                continue
            title = a.get_text(strip=True)
            if url in seen:
                continue
            if (not title or len(title) < 10) and a.find_parent(["h1", "h2", "h3"]):
                title = a.find_parent(["h1", "h2", "h3"]).get_text(strip=True)
            if url in seen or not title or len(title) < 10:
                continue
            if not find_keywords(title):
                continue
            seen.add(url)
            s = get(url, "https://customnews.pk/")
            body = ""
            if s:
                for sel in (".entry-content", ".post-content", "article"):
                    el = s.select_one(sel)
                    if el and len(el.get_text(" ", strip=True)) > 80:
                        body = el.get_text(" ", strip=True)
                        break
            arts.append({
                "title": title, "url": url, "source": "Custom News",
                "date": target, "body": body,
            })
    print(f"  Found {len(arts)} keyword hits")
    return arts


SCRAPERS = {
    "dawn": scrape_dawn,
    "the news": scrape_thenews,
    "express tribune": scrape_tribune,
    "business recorder": scrape_brecorder,
    "jang news english": scrape_jang,
    "custom news": scrape_customnews,
}


WEEKDAY_NAMES = {
    "monday": 0, "mon": 0,
    "tuesday": 1, "tue": 1, "tues": 1,
    "wednesday": 2, "wed": 2,
    "thursday": 3, "thu": 3, "thur": 3, "thurs": 3,
    "friday": 4, "fri": 4,
    "saturday": 5, "sat": 5,
    "sunday": 6, "sun": 6,
}


def parse_datetime_input(raw: str) -> Tuple[str, str]:
    """
    Parse one edition start stamp.
    Preferred: 23-07-2026 14:00  (DD-MM-YYYY HH:MM)
    Also:      23-07-2026, today, yesterday, saturday, 2026-07-23
    Returns: (edition_date YYYY-MM-DD, run_stamp YYYY-MM-DD HH:MM:SS)
    """
    today = date.today()
    raw = (raw or "").strip()
    if not raw:
        now = datetime.now()
        return today.isoformat(), now.strftime("%Y-%m-%d %H:%M:%S")

    # DD-MM-YYYY HH:MM  or  DD-MM-YYYY HH:MM:SS
    m = re.match(
        r"^(\d{2})-(\d{2})-(\d{4})"
        r"(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?$",
        raw,
    )
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        edition = date(y, mo, d)
        if m.group(4) is not None:
            hh, mm = int(m.group(4)), int(m.group(5))
            ss = int(m.group(6) or 0)
            run_dt = datetime(y, mo, d, hh, mm, ss)
        else:
            now = datetime.now()
            run_dt = datetime(y, mo, d, now.hour, now.minute, now.second)
        return edition.isoformat(), run_dt.strftime("%Y-%m-%d %H:%M:%S")

    # YYYY-MM-DD optional time
    m = re.match(
        r"^(\d{4})-(\d{2})-(\d{2})"
        r"(?:\s+(\d{1,2}):(\d{2})(?::(\d{2}))?)?$",
        raw,
    )
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        edition = date(y, mo, d)
        if m.group(4) is not None:
            hh, mm = int(m.group(4)), int(m.group(5))
            ss = int(m.group(6) or 0)
            run_dt = datetime(y, mo, d, hh, mm, ss)
        else:
            now = datetime.now()
            run_dt = datetime(y, mo, d, now.hour, now.minute, now.second)
        return edition.isoformat(), run_dt.strftime("%Y-%m-%d %H:%M:%S")

    low = raw.lower()
    if low == "today":
        now = datetime.now()
        return today.isoformat(), now.strftime("%Y-%m-%d %H:%M:%S")
    if low == "yesterday":
        now = datetime.now()
        ed = today - timedelta(days=1)
        return ed.isoformat(), now.strftime("%Y-%m-%d %H:%M:%S")
    if low in WEEKDAY_NAMES:
        target_wd = WEEKDAY_NAMES[low]
        delta = (today.weekday() - target_wd) % 7
        ed = today - timedelta(days=delta)
        now = datetime.now()
        return ed.isoformat(), now.strftime("%Y-%m-%d %H:%M:%S")

    raise ValueError(
        f"Invalid date/time: {raw!r}\n"
        f"Use: 23-07-2026 14:00   (DD-MM-YYYY HH:MM)"
    )


def ask_datetime(default: str = "") -> Tuple[str, str]:
    """Always ask for one date+time before scraping starts."""
    today = date.today()
    print()
    print("=" * 50)
    print("Pakistan Adverse Media Scraper (portable)")
    print("=" * 50)
    print(f"Today is {today.strftime('%A')} {today.strftime('%d-%m-%Y')}")
    print()
    print("Enter ONE edition date and time, then press Enter.")
    print("  Format:  DD-MM-YYYY HH:MM")
    print("  Example: 23-07-2026 14:00")
    print()
    print("When finished, run again and type the NEXT date/time.")
    print("(Saturday, then Sunday, then Monday on a Monday morning.)")
    print()
    if default:
        print(f"  (Press Enter to use: {default})")
    sys.stdout.flush()
    try:
        raw = input("Date and time> ").strip()
    except EOFError:
        raw = ""
    if not raw:
        raw = default
    if not raw:
        print("No date entered. Using today.")
        return parse_datetime_input("")
    return parse_datetime_input(raw)


def process_articles(
    articles: List[dict], seen: Set[str], stats: dict, run_stamp: str
) -> None:
    for art in articles:
        stats["checked"] += 1
        ok, kws, reason = is_relevant(art["title"], art.get("body", ""))
        if not kws:
            continue
        stats["matched"] += 1
        if not ok:
            stats["filtered"] += 1
            continue
        sid = story_id(art["title"])
        if sid in seen:
            stats["dup"] += 1
            print(f"  [SKIP] dup | {art['source']} | {art['title'][:50]}")
            continue
        pk = primary_kw(kws)
        city = extract_city(art["title"], art.get("body", ""), art.get("url", ""))
        pdf = OUT / pk / art["source"] / art["date"] / f"{safe_name(art['title'])}.pdf"
        save_pdf(
            pdf, art["title"], art["source"], art["url"], art["date"],
            art.get("body", ""), kws,
        )
        append_manifest({
            "record_id": sid,
            "title": art["title"],
            "url": art["url"],
            "keyword": pk,
            "all_keywords": "; ".join(kws),
            "source": art["source"],
            "city": city,
            "date": art["date"],
            "pdf_path": str(pdf.relative_to(OUT)),
            "pdf_link": pdf_hyperlink(pdf),
            "pdf_abs": str(pdf.resolve()),
            "person_names": "",
            "scraped_at": run_stamp,
        })
        seen.add(sid)
        stats["saved"] += 1
        where = city or "-"
        print(f"  [SAVED] {pk} | {art['source']} | {where} | {art['title'][:50]}")


def run_one_day(
    target: str, scrapers: list, seen: Set[str], stats: dict, run_stamp: str
) -> None:
    global TARGET
    TARGET = target
    weekday = datetime.strptime(target, "%Y-%m-%d").strftime("%A")
    print("\n" + "#" * 50)
    print(f"# EDITION DATE: {target} ({weekday})")
    print(f"# START TIME:   {run_stamp}")
    print("#" * 50)

    for name, fn in scrapers:
        try:
            articles = fn(target)
        except Exception as e:
            print(f"  [ERROR] {name}: {e}")
            continue
        process_articles(articles, seen, stats, run_stamp)


def main() -> int:
    global TARGET
    ap = argparse.ArgumentParser(
        description="Pakistan adverse media scraper (portable). "
        "Always prompts for date/time like: 23-07-2026 14:00"
    )
    ap.add_argument(
        "--date",
        default="",
        help='Optional default at prompt, e.g. "23-07-2026 14:00"',
    )
    ap.add_argument(
        "--no-prompt",
        action="store_true",
        help="Skip prompt; requires --date (automation only)",
    )
    ap.add_argument("--sources", nargs="+", help="Limit sources by name")
    ap.add_argument("--fresh", action="store_true", help="Clear downloads + manifest")
    args = ap.parse_args()

    try:
        if args.no_prompt:
            if not args.date:
                print('With --no-prompt you must pass --date "DD-MM-YYYY HH:MM"')
                return 1
            target, run_stamp = parse_datetime_input(args.date)
        else:
            # ALWAYS prompt so you type the date/time before scraping
            target, run_stamp = ask_datetime(default=args.date or "")
    except ValueError as e:
        print(e)
        return 1
    except (EOFError, KeyboardInterrupt):
        print("\nCancelled.")
        return 1

    if args.fresh:
        if OUT.exists():
            shutil.rmtree(OUT)
        clear_manifests()
        print("[fresh] Cleared downloads/ and all manifest*.csv files")

    OUT.mkdir(parents=True, exist_ok=True)
    ensure_manifest_schema(MANIFEST)
    for p in BASE.glob("manifest_*.csv"):
        ensure_manifest_schema(p)
    seen = load_seen()
    if seen:
        print(f"Duplicate memory: {len(seen)} titles from manifest*.csv "
              f"(deleting downloads alone does NOT clear this)")

    print()
    print(f"Edition date: {target}")
    print(f"Start time:   {run_stamp}")
    print(f"Output:       {OUT}")
    print("-" * 50)

    scrapers = list(SCRAPERS.items())
    if args.sources:
        want = {s.lower() for s in args.sources}
        scrapers = [(k, fn) for k, fn in scrapers if k in want]

    stats = {"checked": 0, "matched": 0, "saved": 0, "dup": 0, "filtered": 0}
    run_one_day(target, scrapers, seen, stats, run_stamp)

    print("\n" + "=" * 50)
    print("DONE")
    print(f"  Edition:   {target}")
    print(f"  Started:   {run_stamp}")
    print(f"  Checked:   {stats['checked']}")
    print(f"  Keywords:  {stats['matched']}")
    print(f"  Filtered:  {stats['filtered']}")
    print(f"  Saved:     {stats['saved']}")
    print(f"  Dupes:     {stats['dup']}")
    print(f"  Manifest:  {MANIFEST}")
    print("=" * 50)
    print()
    print("Re-run the program and enter the NEXT date/time to continue.")
    return 0


if __name__ == "__main__":
    sys.exit(main())